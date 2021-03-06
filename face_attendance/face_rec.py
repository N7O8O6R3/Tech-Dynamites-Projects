import face_recognition as fr
import os
import cv2
import face_recognition
import numpy as np
import time
import shelve
import openpyxl as xl



def unknown_image_encoded(img):
    """
    encode a face given the file name
    """
    face = fr.load_image_file("faces/" + img)
    encoding = fr.face_encodings(face)[0]

    return encoding


def classify_face(im):
    """
    will find all of the faces in a given image and label
    them if it knows what they are

    :param im: str of file path
    :return: list of face names
    """
    #get_encoded_faces()
    faces = shelve.open('trainingData.yml')
    #faces = faces1.read()
    #print(faces)
    faces_encoded = list(faces.values())
    known_face_names = list(faces.keys())

    img = cv2.imread(im, 1)
    #img = cv2.resize(img, (0, 0), fx=0.5, fy=0.5)
    #img = img[:,:,::-1]
 
    face_locations = face_recognition.face_locations(img)
    unknown_face_encodings = face_recognition.face_encodings(img, face_locations)

    face_names = []
    for face_encoding in unknown_face_encodings:
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(faces_encoded, face_encoding)
        name = "Unknown"
        #print("face_names",face_names)
        #print("faces_encoded",faces_encoded)
        #print("known_fac_names:",known_face_names)

        # use the known face with the smallest distance to the new face
        face_distances = face_recognition.face_distance(faces_encoded, face_encoding)
        best_match_index = np.argmin(face_distances)
        if matches[best_match_index]:
            name = known_face_names[best_match_index]

        face_names.append(name)

        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Draw a box around the face
            cv2.rectangle(img, (left-20, top-20), (right+20, bottom+20), (255, 0, 0), 2)

            # Draw a label with a name below the face
            cv2.rectangle(img, (left-20, bottom -15), (right+20, bottom+20), (255, 0, 0), cv2.FILLED)
            font = cv2.FONT_HERSHEY_COMPLEX_SMALL
            cv2.putText(img, name, (left -20, bottom + 15), font, 1.0, (255, 255, 255), 2)


    # Display the resulting image
    while True:

        cv2.imshow('Video', img)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            return face_names 



attendance = classify_face("test.jpg")

try:

	wb = xl.load_workbook('attendancef.xlsx')
	ws= wb.active
	#ws . append(attendance)
	#sec = time.time()
	local = time.strftime("%d/%m/%Y")
	col = ws.max_column
	need =ws.cell(row=1,column=col)
	if need.value != local:
		col +=1
		c1 = ws.cell(row = 1, column = col)
		c1.value = local
	else:
		ws.delete_cols(col)
		c1 = ws.cell(row = 1, column = col)
		c1.value = local
	rcount = 2
	for i in attendance :
		c1 = ws.cell(row = rcount, column = col)
		c1.value = i
		rcount +=1
	wb.save('attendancef.xlsx')
except FileNotFoundError:
	wb = xl.Workbook()
	ws = wb.active
	ws.title="attendance"
	#sec = time.time()
	local = time.strftime("%d/%m/%Y")
	c1 = ws.cell(row = 1, column = 1)
	c1.value = local
	rcount = 2
	for i in attendance :
		c1 = ws.cell(row = rcount, column = 1)
		c1.value = i
		rcount +=1
	wb.save('attendancef.xlsx')
